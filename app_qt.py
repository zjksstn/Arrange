# -*- coding: utf-8 -*-
"""
瑞士制编排 —— Qt 新界面(含组别)
==================================
三栏式仿"云蛇":
  上    工具栏(切换/新建赛事 · 组别 · 撤销本轮 · 编排下一轮)
  左    赛事文件区(二级树:赛事>组别 + 新建/打开/删除) + 打印报表区
  主区  轮次选项卡 + 可视化对阵卡片(红胜/和/黑胜直接登分)
  底    状态栏(赛事 · 组别 · 轮次 · 未登分台号 · 完成度)

数据模型:Event(赛事)含多个 Group(组别),每组是一套独立瑞士制。
切到哪个组,主区就重建成那个组的登分界面;全程自动存盘到 events/*.json。

依赖:tournament.py / pairing.py / print_pdf.py
运行:venv\\Scripts\\python.exe app_qt.py
"""
import os
import sys

from PySide6.QtWidgets import (
    QApplication, QMainWindow, QWidget, QFrame, QLabel, QPushButton,
    QVBoxLayout, QHBoxLayout, QGridLayout, QTabWidget, QScrollArea,
    QTreeWidget, QTreeWidgetItem, QToolBar, QToolButton, QDialog, QFormLayout,
    QLineEdit, QSpinBox, QPlainTextEdit, QDialogButtonBox, QFileDialog,
    QMessageBox, QMenu, QInputDialog, QStackedWidget, QTableWidget,
    QTableWidgetItem, QAbstractItemView, QHeaderView, QCheckBox, QComboBox,
)
from PySide6.QtGui import QAction
from PySide6.QtCore import Qt

import tournament

DEMO_FOLDER = os.path.join(os.path.dirname(__file__), "mathdate", "2013nzjkzgxqs")
RECENT_N = 4               # 左栏最多显示最近几个赛事
# 开发模式:设环境变量 ARRANGE_DEV=1 时显示「随机结果(测试)」按钮;正式版默认隐藏
DEV_MODE = os.environ.get("ARRANGE_DEV", "") == "1"


STYLE = """
QMainWindow, QWidget#central { background: #f1efe8; }
QToolBar { background: #ffffff; border-bottom: 1px solid #d7d2c4; padding: 4px; spacing: 2px; }
QToolButton { padding: 5px 12px; border-radius: 6px; color: #2c2c2a; font-size: 13px; }
QToolButton:hover { background: #ece9e0; }
QToolButton:disabled { color: #b4b2a9; }
QWidget#left { background: #f6f4ee; border-right: 1px solid #d7d2c4; }
QLabel[role="grouptitle"] { color: #888780; font-size: 12px; }
QPushButton[role="report"] {
    text-align: left; padding: 6px 10px; font-size: 13px;
    border: 1px solid #d7d2c4; border-radius: 6px; background: #ffffff; color: #2c2c2a;
}
QPushButton[role="report"]:hover { background: #ece9e0; }
QPushButton[role="filebtn"] {
    padding: 4px 0; font-size: 12px; border: 1px solid #d7d2c4;
    border-radius: 6px; background: #ffffff; color: #2c2c2a;
}
QPushButton[role="filebtn"]:hover { background: #ece9e0; }
QTreeWidget { border: 1px solid #d7d2c4; border-radius: 6px; background: #ffffff; font-size: 13px; }
QTabWidget::pane { border: 1px solid #d7d2c4; background: #f4f2ec; top: -1px; }
QTabBar::tab {
    background: #ece9e0; color: #5f5e5a; padding: 6px 16px; margin-right: 2px;
    border: 1px solid #d7d2c4; border-bottom: none;
    border-top-left-radius: 8px; border-top-right-radius: 8px;
}
QTabBar::tab:selected { background: #ffffff; color: #2c2c2a; font-weight: 600; }

QFrame#card { background: #ffffff; border: 1px solid #d7d2c4; border-radius: 12px; }
QFrame#card[scored="true"] { border: 2px solid #4a9e6f; }
QFrame#card[bye="true"] { background: #f6f4ee; border: 1px dashed #b4b2a9; }
QLabel#tai { font-size: 15px; font-weight: 600; color: #2c2c2a; }
QLabel#status { font-size: 12px; color: #888780; }
QLabel#status[scored="true"] { color: #4a9e6f; }
QLabel[role="name"] { font-size: 15px; color: #2c2c2a; }
QLabel[role="team"] { font-size: 12px; color: #888780; }
QLabel#vs { color: #b4b2a9; font-size: 12px; }
QPushButton[role="res"] {
    border: 1px solid #d7d2c4; border-radius: 6px;
    padding: 6px 0; font-size: 13px; background: #f7f6f2; color: #2c2c2a;
}
QPushButton[role="res"]:hover { background: #eceae3; }
QPushButton[role="res"]:disabled { color: #b4b2a9; background: #f1efe8; }
QPushButton[on="true"] { background: #e1f3ea; border: 1px solid #4a9e6f; color: #2b6e4f; }
QStatusBar { background: #ffffff; border-top: 1px solid #d7d2c4; color: #5f5e5a; }
QStatusBar QLabel { font-size: 12px; padding: 0 8px; }
"""


class BoardCard(QFrame):
    RESULTS = ["红胜", "和", "黑胜"]

    def __init__(self, board, on_scored=None, locked=False):
        super().__init__()
        self.board = board
        self.on_scored = on_scored
        self.locked = locked
        self.setObjectName("card")
        self.setProperty("scored", "false")
        self.setProperty("bye", "true" if board.is_bye else "false")

        root = QVBoxLayout(self)
        root.setContentsMargins(14, 12, 14, 12); root.setSpacing(10)

        top = QHBoxLayout()
        tai_lbl = QLabel("第 %d 台" % board.tai); tai_lbl.setObjectName("tai")
        self.status = QLabel(); self.status.setObjectName("status")
        self.status.setProperty("scored", "false")
        top.addWidget(tai_lbl); top.addStretch(); top.addWidget(self.status)
        root.addLayout(top)

        players = QHBoxLayout()
        players.addLayout(self._player(board.red, "#d85a30"))
        if board.is_bye:
            bye = QLabel("— 轮空 —"); bye.setProperty("role", "team")
            bye.setAlignment(Qt.AlignRight)
            players.addStretch(); players.addWidget(bye)
        else:
            vs = QLabel("VS"); vs.setObjectName("vs"); vs.setAlignment(Qt.AlignCenter)
            players.addWidget(vs)
            players.addLayout(self._player(board.black, "#2c2c2a", right=True))
        root.addLayout(players)

        if board.is_bye:
            self.status.setText("轮空 · 自动胜")
            self.status.setProperty("scored", "true")
            self.setProperty("scored", "true")
        else:
            self.status.setText("待登分")
            row = QHBoxLayout(); row.setSpacing(6)
            self.buttons = []
            for i, text in enumerate(self.RESULTS):
                btn = QPushButton(text)
                btn.setProperty("role", "res"); btn.setProperty("on", "false")
                btn.clicked.connect(lambda checked=False, idx=i: self.on_result(idx))
                row.addWidget(btn); self.buttons.append(btn)
            root.addLayout(row)
            if board.result:
                self._apply_result(board.result - 1)
            if locked:
                for btn in self.buttons:
                    btn.setEnabled(False)
                if board.result:
                    self.status.setText("%s · 已锁定" % self.RESULTS[board.result - 1])

    def _player(self, p, dot_color, right=False):
        box = QVBoxLayout(); box.setSpacing(2)
        name = QLabel('<span style="color:%s">●</span> %s' % (dot_color, p.name))
        name.setProperty("role", "name")
        team = QLabel("%s · %d分" % (p.team, p.frac))
        team.setProperty("role", "team")
        if right:
            name.setAlignment(Qt.AlignRight); team.setAlignment(Qt.AlignRight)
        box.addWidget(name); box.addWidget(team)
        return box

    def on_result(self, idx):
        if self.locked:
            return
        self.board.result = idx + 1
        self._apply_result(idx)
        if self.on_scored:
            self.on_scored()

    def _apply_result(self, idx):
        for i, btn in enumerate(self.buttons):
            btn.setProperty("on", "true" if i == idx else "false")
            self._repolish(btn)
        self.setProperty("scored", "true")
        self.status.setText("已登分  %s" % self.RESULTS[idx])
        self.status.setProperty("scored", "true")
        self._repolish(self); self._repolish(self.status)

    @staticmethod
    def _repolish(w):
        w.style().unpolish(w); w.style().polish(w)


def make_round_view(boards, on_scored=None, locked=False):
    container = QWidget()
    grid = QGridLayout(container)
    grid.setContentsMargins(12, 12, 12, 12); grid.setSpacing(10)
    if not boards:
        tip = QLabel("本轮尚未编排"); tip.setStyleSheet("color:#888780;font-size:14px;")
        grid.addWidget(tip, 0, 0)
    else:
        ncols = 2
        for i, b in enumerate(boards):
            grid.addWidget(BoardCard(b, on_scored, locked), i // ncols, i % ncols)
        grid.setRowStretch(len(boards), 1)
    scroll = QScrollArea(); scroll.setWidgetResizable(True)
    scroll.setWidget(container); scroll.setFrameShape(QFrame.NoFrame)
    return scroll


def round_label(n):
    return "首轮" if n == 1 else "第 %d 轮" % n


def open_pdf_or_inform(parent, out):
    """打开 PDF 供查看/打印;若系统未能自动打开(无默认 PDF 程序等),
    则提示文件已保存的位置,让用户手动打开打印 —— 文件已存档,绝不丢失。"""
    import print_pdf
    try:
        print_pdf.open_in_viewer(out)
    except Exception:
        QMessageBox.information(parent, "PDF 已生成",
                               "PDF 已保存到:\n%s\n\n系统未能自动打开它,请手动打开该文件进行打印。" % out)


class _RosterDialogBase(QDialog):
    """带"选手名单(每行一个,可从 txt 导入)"的对话框基类。"""
    def _add_roster_row(self, form):
        self.names_edit = QPlainTextEdit()
        self.names_edit.setPlaceholderText("每行一个选手姓名")
        form.addRow("选手名单", self.names_edit)
        imp = QPushButton("从 txt 文件导入名单…")
        imp.clicked.connect(self._import_names)
        form.addRow("", imp)

    def _import_names(self):
        fn, _ = QFileDialog.getOpenFileName(self, "选择名单文件", "",
                                            "文本文件 (*.txt);;所有文件 (*)")
        if fn:
            with open(fn, encoding="utf-8") as f:
                self.names_edit.setPlainText(f.read())

    def _names(self):
        return [ln.strip() for ln in self.names_edit.toPlainText().splitlines() if ln.strip()]


class NewEventDialog(_RosterDialogBase):
    """新建赛事:赛事名 / 裁判长 / 首个组别名 / 该组轮数 / 该组名单。"""
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("新建赛事")
        self.resize(380, 470)
        form = QFormLayout(self)
        self.name_edit = QLineEdit()
        self.judge_edit = QLineEdit()
        self.group_edit = QLineEdit("公开组")
        self.rounds_spin = QSpinBox(); self.rounds_spin.setRange(1, 30); self.rounds_spin.setValue(7)
        form.addRow("赛事名称", self.name_edit)
        form.addRow("裁判长", self.judge_edit)
        form.addRow("首个组别", self.group_edit)
        form.addRow("该组总轮数", self.rounds_spin)
        self._add_roster_row(form)
        box = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        box.accepted.connect(self.accept); box.rejected.connect(self.reject)
        form.addRow(box)

    def data(self):
        return (self.name_edit.text().strip(), self.judge_edit.text().strip(),
                self.group_edit.text().strip(), self.rounds_spin.value(), self._names())


class NewGroupDialog(_RosterDialogBase):
    """新建组别:组别名 / 轮数 / 名单(加入当前赛事)。"""
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("新建组别")
        self.resize(380, 420)
        form = QFormLayout(self)
        self.group_edit = QLineEdit()
        self.rounds_spin = QSpinBox(); self.rounds_spin.setRange(1, 30); self.rounds_spin.setValue(7)
        form.addRow("组别名称", self.group_edit)
        form.addRow("总轮数", self.rounds_spin)
        self._add_roster_row(form)
        box = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        box.accepted.connect(self.accept); box.rejected.connect(self.reject)
        form.addRow(box)

    def data(self):
        return (self.group_edit.text().strip(), self.rounds_spin.value(), self._names())


class RosterTable(QTableWidget):
    """报名阶段的花名册表格:可直接编辑姓名/性别/身份证/电话/等级分/K值。"""
    COLS = [("编号", "number", True), ("姓名", "name", False), ("性别", "gender", False),
            ("身份证号", "idcard", False), ("电话", "phone", False),
            ("等级分", "rating", False), ("K值", "kvalue", False)]

    def __init__(self, group, on_edit, readonly=False):
        super().__init__()
        self.group = group
        self.on_edit = on_edit
        self.readonly = readonly
        self._filling = False
        self.setColumnCount(len(self.COLS))
        self.setHorizontalHeaderLabels([c[0] for c in self.COLS])
        self.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        if readonly:
            self.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.cellChanged.connect(self._on_cell_changed)
        self.reload()

    def reload(self):
        self._filling = True
        self.setRowCount(len(self.group.players))
        for r, p in enumerate(self.group.players):
            for c, (label, attr, ro) in enumerate(self.COLS):
                val = getattr(p, attr)
                text = "" if val is None else str(val)
                if self.readonly and c == 1 and p.withdrawn:   # 只读视图:退赛者姓名标注
                    text += "(退赛)"
                it = QTableWidgetItem(text)
                if ro or self.readonly:
                    it.setFlags(it.flags() & ~Qt.ItemIsEditable)
                self.setItem(r, c, it)
        self._filling = False

    def _on_cell_changed(self, row, col):
        if self._filling or self.readonly:
            return
        p = self.group.players[row]
        _, attr, ro = self.COLS[col]
        if ro:
            return
        text = self.item(row, col).text().strip()
        if attr == "rating":
            try:
                p.rating = float(text or 0)
            except ValueError:
                return
        elif attr == "kvalue":
            try:
                p.kvalue = int(float(text or 0))
            except ValueError:
                return
        else:
            setattr(p, attr, text)
        self.on_edit()

    def selected_numbers(self):
        rows = sorted({i.row() for i in self.selectedIndexes()})
        return [self.group.players[r].number for r in rows]


class SelectPlayersDialog(QDialog):
    """勾选选手对话框(退赛/复赛共用):复选框 + 编号/姓名/团队/积分。"""
    def __init__(self, title, players, parent=None):
        super().__init__(parent)
        self.setWindowTitle(title)
        self.resize(440, 480)
        self.players = list(players)
        lay = QVBoxLayout(self)
        self.table = QTableWidget()
        cols = ["选择", "编号", "姓名", "团队", "积分"]
        self.table.setColumnCount(len(cols)); self.table.setHorizontalHeaderLabels(cols)
        self.table.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.table.setRowCount(len(self.players))
        for r, p in enumerate(self.players):
            chk = QTableWidgetItem()
            chk.setFlags(Qt.ItemIsUserCheckable | Qt.ItemIsEnabled)
            chk.setCheckState(Qt.Unchecked)
            self.table.setItem(r, 0, chk)
            for c, v in enumerate([p.number, p.name, p.team, p.frac], start=1):
                self.table.setItem(r, c, QTableWidgetItem(str(v)))
        lay.addWidget(self.table)
        box = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        box.accepted.connect(self.accept); box.rejected.connect(self.reject)
        lay.addWidget(box)

    def selected_numbers(self):
        return [self.players[r].number for r in range(self.table.rowCount())
                if self.table.item(r, 0).checkState() == Qt.Checked]


class StandingsDialog(QDialog):
    """最终成绩表(只读)+ 导出/打印 PDF。列序:姓名…名次。"""
    COLS = ["姓名", "积分", "对手分", "后手", "犯规", "签号", "备注", "团队", "名次"]

    def __init__(self, group, match_name="", judge="", parent=None):
        super().__init__(parent)
        self.group = group
        self.match_name = match_name or group.name
        self.judge = judge
        self.setWindowTitle("最终成绩 — %s" % self.match_name)
        self.resize(560, 540)
        lay = QVBoxLayout(self)
        table = QTableWidget()
        table.setColumnCount(len(self.COLS)); table.setHorizontalHeaderLabels(self.COLS)
        table.setEditTriggers(QAbstractItemView.NoEditTriggers)
        table.setSelectionBehavior(QAbstractItemView.SelectRows)
        table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        st = group.standings()
        table.setRowCount(len(st))
        for r, (rank, p, note, hlcols) in enumerate(st):
            for c, v in enumerate(self._row(rank, p, note)):
                it = QTableWidgetItem(str(v))
                if self.COLS[c] in hlcols:          # 靠后手/犯规/备注破同分 → 标红加粗
                    it.setForeground(Qt.GlobalColor.red)
                    f = it.font(); f.setBold(True); it.setFont(f)
                table.setItem(r, c, it)
        lay.addWidget(table)

        box = QDialogButtonBox(QDialogButtonBox.Close)
        export_btn = QPushButton("导出 / 打印 PDF")
        export_btn.clicked.connect(self.on_export)
        box.addButton(export_btn, QDialogButtonBox.ActionRole)
        box.rejected.connect(self.reject)
        lay.addWidget(box)

    @staticmethod
    def _row(rank, p, note):
        name = p.name + ("(退赛)" if p.withdrawn else "")
        return [name, p.frac, p.competfrac, p.backcount, p.foulcount, p.number, note, p.team, rank]

    def on_export(self):
        import print_pdf, tournament
        rows = [[str(x) for x in self._row(rank, p, note)]
                for rank, p, note, _ in self.group.standings()]
        out = os.path.join(tournament.events_dir(),
                           "%s_最终成绩表.pdf" % tournament.safe_filename(self.match_name))
        try:
            print_pdf.generate_standings_pdf(rows, self.match_name, out, judge=self.judge)
        except Exception as e:
            QMessageBox.critical(self, "导出失败", str(e)); return
        open_pdf_or_inform(self, out)


class EditInfoDialog(QDialog):
    """编辑赛事信息:赛事名 / 裁判长(赛事级) + 当前组别名 / 总轮数(组级)。"""
    def __init__(self, event_name, judge, group_name, rounds, parent=None):
        super().__init__(parent)
        self.setWindowTitle("编辑赛事信息")
        self.resize(360, 240)
        form = QFormLayout(self)
        self.name_edit = QLineEdit(event_name)
        self.judge_edit = QLineEdit(judge)
        self.group_edit = QLineEdit(group_name)
        self.rounds_spin = QSpinBox(); self.rounds_spin.setRange(1, 30); self.rounds_spin.setValue(rounds)
        form.addRow("赛事名称", self.name_edit)
        form.addRow("裁判长", self.judge_edit)
        form.addRow("当前组别名称", self.group_edit)
        form.addRow("当前组总轮数", self.rounds_spin)
        box = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        box.accepted.connect(self.accept); box.rejected.connect(self.reject)
        form.addRow(box)

    def data(self):
        return (self.name_edit.text().strip(), self.judge_edit.text().strip(),
                self.group_edit.text().strip(), self.rounds_spin.value())


class ExcelImportDialog(QDialog):
    """从 Excel(.xlsx)导入名单:选工作表 → 把各字段对应到某一列 → 预览 → 导入。"""
    FIELDS = [("姓名", "name", True), ("性别", "gender", False), ("团队", "team", False),
              ("身份证号", "idcard", False), ("电话", "phone", False),
              ("等级分", "rating", False), ("K值", "kvalue", False)]

    def __init__(self, path, parent=None):
        super().__init__(parent)
        self.setWindowTitle("从 Excel 导入名单")
        self.resize(640, 560)
        import openpyxl
        self.wb = openpyxl.load_workbook(path, data_only=True)
        lay = QVBoxLayout(self)

        top = QHBoxLayout()
        top.addWidget(QLabel("工作表:"))
        self.sheet_combo = QComboBox(); self.sheet_combo.addItems(self.wb.sheetnames)
        self.sheet_combo.currentIndexChanged.connect(self._on_sheet_changed)
        top.addWidget(self.sheet_combo)
        self.header_cb = QCheckBox("首行为标题(跳过)"); self.header_cb.setChecked(True)
        self.header_cb.stateChanged.connect(self._refresh_preview)
        top.addWidget(self.header_cb); top.addStretch()
        lay.addLayout(top)

        form = QFormLayout()
        self.combos = {}
        for label, attr, req in self.FIELDS:
            cb = QComboBox(); cb.currentIndexChanged.connect(self._refresh_preview)
            self.combos[attr] = cb
            form.addRow(label + ("（必填）" if req else "（可选）"), cb)
        lay.addLayout(form)

        lay.addWidget(QLabel("预览(前 8 行):"))
        self.preview = QTableWidget()
        self.preview.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.preview.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        lay.addWidget(self.preview)

        box = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        box.accepted.connect(self.accept); box.rejected.connect(self.reject)
        lay.addWidget(box)
        self._on_sheet_changed()

    def _sheet(self):
        return self.wb[self.sheet_combo.currentText()]

    def _on_sheet_changed(self):
        n = self._sheet().max_column or 1
        opts = ["—不导入—"] + ["第%d列" % (i + 1) for i in range(n)]
        for attr, cb in self.combos.items():
            cb.blockSignals(True); cb.clear(); cb.addItems(opts); cb.blockSignals(False)
        self.combos["name"].setCurrentIndex(1)       # 默认 姓名 → 第1列
        self._refresh_preview()

    def _col_index(self, attr):
        i = self.combos[attr].currentIndex()
        return None if i <= 0 else i - 1             # 0-based;0=不导入

    def _rows(self):
        rows = list(self._sheet().iter_rows(values_only=True))
        if self.header_cb.isChecked() and rows:
            rows = rows[1:]
        return rows

    def _cell(self, row, attr):
        ci = self._col_index(attr)
        if ci is None or ci >= len(row):
            return ""
        v = row[ci]
        return "" if v is None else str(v).strip()

    def _refresh_preview(self):
        rows = self._rows()[:8]
        labels = [lb for lb, attr, req in self.FIELDS]
        self.preview.setColumnCount(len(labels)); self.preview.setHorizontalHeaderLabels(labels)
        self.preview.setRowCount(len(rows))
        for r, row in enumerate(rows):
            for c, (lb, attr, req) in enumerate(self.FIELDS):
                self.preview.setItem(r, c, QTableWidgetItem(self._cell(row, attr)))

    def players(self):
        """返回 [dict(name, [gender, team, idcard, phone, rating, kvalue]), ...](姓名非空)。"""
        out = []
        for row in self._rows():
            name = self._cell(row, "name")
            if not name:
                continue
            d = {"name": name}
            for label, attr, req in self.FIELDS:
                if attr == "name":
                    continue
                val = self._cell(row, attr)
                if attr == "rating":
                    try: d["rating"] = float(val) if val else 0.0
                    except ValueError: d["rating"] = 0.0
                elif attr == "kvalue":
                    try: d["kvalue"] = int(float(val)) if val else 0
                    except ValueError: d["kvalue"] = 0
                elif val:
                    d[attr] = val
            out.append(d)
        return out


class LeftPanel(QWidget):
    REPORTS = ["本轮对阵表", "总成绩表", "荣誉证书", "选手花名册", "台号卡", "记分卡"]

    def __init__(self, cb):
        super().__init__()
        self._cb = cb
        self._loading = False          # 重建树时抑制选中回调,防止递归
        self.setObjectName("left"); self.setFixedWidth(200)
        lay = QVBoxLayout(self); lay.setContentsMargins(10, 10, 10, 10); lay.setSpacing(6)

        lay.addWidget(self._title("当前赛事(赛事 › 组别)"))
        self.tree = QTreeWidget(); self.tree.setHeaderHidden(True); self.tree.setFixedHeight(180)
        self.tree.currentItemChanged.connect(self._on_current_changed)   # 选中即切换
        self.tree.setContextMenuPolicy(Qt.CustomContextMenu)             # 右键菜单
        self.tree.customContextMenuRequested.connect(self._on_context_menu)
        lay.addWidget(self.tree)

        edit_btn = QPushButton("编辑赛事信息"); edit_btn.setProperty("role", "filebtn")
        edit_btn.clicked.connect(lambda: cb["edit_info"]())
        lay.addWidget(edit_btn)
        gb = QHBoxLayout(); gb.setSpacing(4)
        for text, key in [("＋ 新建组别", "new_group"), ("－ 删除组别", "del_group")]:
            b = QPushButton(text); b.setProperty("role", "filebtn")
            b.clicked.connect(lambda checked=False, k=key: cb[k]())
            gb.addWidget(b)
        lay.addLayout(gb)

        lay.addSpacing(6); lay.addWidget(self._title("打印报表"))
        self.report_btns = {}
        for name in self.REPORTS:
            b = QPushButton(name); b.setProperty("role", "report")
            b.clicked.connect(lambda checked=False, n=name: cb["report"](n))
            lay.addWidget(b)
            self.report_btns[name] = b
        lay.addStretch()

    def update_reports(self, started, finished):
        """比赛中:对阵表可用、成绩表灰;已结束:反之。"""
        self.report_btns["本轮对阵表"].setEnabled(started and not finished)
        self.report_btns["总成绩表"].setEnabled(finished)
        self.report_btns["荣誉证书"].setEnabled(finished)

    def set_tree(self, events, cur_path, cur_gi):
        self._loading = True               # 程序化重建期间不触发切换
        self.tree.clear()
        for name, path, gnames in events:
            ev = QTreeWidgetItem([name]); ev.setData(0, Qt.UserRole, (path, None))
            self.tree.addTopLevelItem(ev)
            for gi, gn in enumerate(gnames):
                g = QTreeWidgetItem([gn]); g.setData(0, Qt.UserRole, (path, gi))
                ev.addChild(g)
                if path == cur_path and gi == cur_gi:
                    self.tree.setCurrentItem(g)
            ev.setExpanded(True)
        self._loading = False

    def _on_current_changed(self, cur, prev):
        if self._loading or cur is None:
            return
        self._cb["open"]()                 # 选中赛事/组别 → 自动切换过去

    def _on_context_menu(self, pos):
        item = self.tree.itemAt(pos)
        if not item:
            return
        self.tree.setCurrentItem(item)     # 先切换到右键的目标
        _, gi = item.data(0, Qt.UserRole)
        menu = QMenu(self)
        if gi is None:                     # 赛事节点
            menu.addAction("重命名赛事", lambda: self._cb["rename_event"]())
            menu.addAction("删除赛事", lambda: self._cb["del_event"]())
        else:                              # 组别节点
            menu.addAction("重命名组别", lambda: self._cb["rename_group"]())
            menu.addAction("删除组别", lambda: self._cb["del_group"]())
        menu.exec(self.tree.viewport().mapToGlobal(pos))

    def selected_target(self):
        it = self.tree.currentItem()
        return it.data(0, Qt.UserRole) if it else None

    @staticmethod
    def _title(text):
        lbl = QLabel(text); lbl.setProperty("role", "grouptitle")
        return lbl


class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.resize(1020, 660)
        self.cur_event = None        # 注意:不能用 self.cur_event,会覆盖 QWidget.event() 导致崩溃
        self.current_path = None
        self.gi = 0

        tb = QToolBar(); tb.setMovable(False); self.addToolBar(tb)

        # 「新建赛事 ▾」下拉:新建 / 删除当前赛事
        new_btn = QToolButton(); new_btn.setText("新建赛事 ▾")
        new_btn.setPopupMode(QToolButton.InstantPopup)
        new_menu = QMenu(self)
        new_menu.addAction("新建赛事…", lambda: self.on_new_event())
        new_menu.addAction("删除当前赛事", lambda: self.on_del_event(self.current_path))
        new_btn.setMenu(new_menu); tb.addWidget(new_btn)

        # 「切换赛事 ▾」下拉:列出全部赛事,点击切换
        switch_btn = QToolButton(); switch_btn.setText("切换赛事 ▾")
        switch_btn.setPopupMode(QToolButton.InstantPopup)
        self.switch_menu = QMenu(self)
        self.switch_menu.aboutToShow.connect(self._populate_switch_menu)
        switch_btn.setMenu(self.switch_menu); tb.addWidget(switch_btn)

        tools = ["花名册", "生成秩序册"] + (["随机结果(测试)"] if DEV_MODE else [])
        for text in tools:
            act = QAction(text, self)
            act.triggered.connect(lambda checked=False, t=text: self.on_tool(t))
            tb.addAction(act)
        tb.addSeparator()
        self.undo_act = QAction("◀ 撤销本轮", self)
        self.undo_act.triggered.connect(self.on_undo); tb.addAction(self.undo_act)
        self.next_act = QAction("编排下一轮 ▶", self)
        self.next_act.triggered.connect(self.on_next_round); tb.addAction(self.next_act)

        self.left = LeftPanel({"open": self.on_open, "edit_info": self.on_edit_info,
                               "del_event": self.on_del_event, "new_group": self.on_new_group,
                               "del_group": self.on_del_group, "report": self.on_report,
                               "rename_event": self.on_rename_event,
                               "rename_group": self.on_rename_group})
        # 主区用堆叠:报名阶段显示花名册,开赛后显示轮次选项卡
        self.tabs = QTabWidget()
        self.roster_holder = QWidget()
        self.roster_layout = QVBoxLayout(self.roster_holder)
        self.roster_layout.setContentsMargins(0, 0, 0, 0)
        self.stack = QStackedWidget()
        self.stack.addWidget(self.tabs)            # index 0:对阵
        self.stack.addWidget(self.roster_holder)   # index 1:花名册
        central = QWidget(); central.setObjectName("central")
        h = QHBoxLayout(central); h.setContentsMargins(0, 0, 0, 0); h.setSpacing(0)
        h.addWidget(self.left); h.addWidget(self.stack)
        self.setCentralWidget(central)

        sb = self.statusBar()
        self.cur_event_lbl = QLabel(); sb.addWidget(self.cur_event_lbl)
        self.hint_lbl = QLabel(); self.hint_lbl.setStyleSheet("color:#b06a00;"); sb.addWidget(self.hint_lbl)
        self.done_lbl = QLabel(); sb.addPermanentWidget(self.done_lbl)

        self.load_startup()

    @property
    def group(self):
        return self.cur_event.groups[self.gi]

    # ── 装载 / 存盘 ───────────────────────────────────────────────
    def load_startup(self):
        events = tournament.list_events()
        if events:
            self.set_active(tournament.load_event(events[0][1]), events[0][1], 0)
        else:
            e = tournament.load_demo_event(DEMO_FOLDER)
            self.set_active(e, tournament.event_path(e.name), 0)

    def set_active(self, event, path, gi):
        self.cur_event = event
        self.current_path = path
        self.gi = max(0, min(gi, len(event.groups) - 1))
        self.setWindowTitle("瑞士制编排 — %s / %s" % (event.name, self.group.name))
        if self.group.started:                   # 已开赛 → 对阵选项卡
            self.build_all_tabs()
            self.stack.setCurrentWidget(self.tabs)
        else:                                    # 报名阶段 → 花名册
            self.build_roster_view()
            self.stack.setCurrentWidget(self.roster_holder)
        self.refresh_status()
        self.persist()
        self.refresh_tree()

    def build_roster_view(self):
        """报名阶段:可编辑花名册 + 抽签开赛;已开赛:只读花名册 + 退赛/复赛/返回对阵。"""
        while self.roster_layout.count():
            w = self.roster_layout.takeAt(0).widget()
            if w:
                w.deleteLater()
        started = self.group.started
        self.roster_table = RosterTable(self.group, self.after_roster_edit, readonly=started)
        bar = QHBoxLayout()
        if not started:                              # 报名阶段:增删导入
            for text, slot in [("添加一行", self.on_roster_add), ("删除选中", self.on_roster_del),
                               ("从 txt 导入", self.on_roster_import_txt),
                               ("从 Excel 导入", self.on_roster_import_excel)]:
                b = QPushButton(text); b.setProperty("role", "filebtn"); b.clicked.connect(slot)
                bar.addWidget(b)
            bar.addStretch()
            # 抽签附加条件(仅作用首轮)
            self.cb_same_team = QCheckBox("同团队首轮不相遇")
            self.cb_lower_red = QCheckBox("首轮序号小者持红")
            self.cb_same_team.setChecked(self.group.opt_same_team)
            self.cb_lower_red.setChecked(self.group.opt_lower_red)
            bar.addWidget(self.cb_same_team); bar.addWidget(self.cb_lower_red)
            start_btn = QPushButton("抽签编排首轮 ▶")
            start_btn.setProperty("role", "report"); start_btn.clicked.connect(self.on_draw_lots)
            bar.addWidget(start_btn)
            tip = "　报名阶段:录入名单 → 勾选抽签条件(仅作用首轮)→ 点「抽签编排首轮」开赛"
        else:                                        # 已开赛:退赛/复赛
            for text, slot in [("退赛…", self.on_withdraw), ("复赛…", self.on_reenter)]:
                b = QPushButton(text); b.setProperty("role", "filebtn"); b.clicked.connect(slot)
                bar.addWidget(b)
            bar.addStretch()
            back_btn = QPushButton("← 返回对阵")
            back_btn.setProperty("role", "report")
            back_btn.clicked.connect(lambda: self.stack.setCurrentWidget(self.tabs))
            bar.addWidget(back_btn)
            tip = "　花名册(已开赛,只读)。退赛/复赛后,当前轮将自动重排(请在未发布对阵表前操作)。"
        top = QWidget(); top.setLayout(bar)
        self.roster_layout.addWidget(QLabel(tip))
        self.roster_layout.addWidget(self.roster_table)
        self.roster_layout.addWidget(top)

    def on_roster(self):
        """工具栏「花名册」:切到花名册页(报名阶段可编辑,已开赛只读+退赛/复赛)。"""
        self.build_roster_view()
        self.stack.setCurrentWidget(self.roster_holder)

    def on_withdraw(self):
        g = self.group
        if g.finished:
            return
        active = [p for p in g.players if not p.withdrawn]
        dlg = SelectPlayersDialog("退赛 — 勾选要退赛的选手", active, self)
        if dlg.exec() != QDialog.Accepted:
            return
        nums = dlg.selected_numbers()
        if not nums:
            return
        if QMessageBox.question(self, "退赛确认",
                                "确定让选中的 %d 名选手退赛?\n当前轮将自动重排,本轮已录入的成绩将作废。" % len(nums),
                                QMessageBox.Yes | QMessageBox.No, QMessageBox.No) != QMessageBox.Yes:
            return
        g.withdraw(nums)
        self._after_roster_change("已退赛 %d 人,当前轮已重排" % len(nums))

    def on_reenter(self):
        g = self.group
        if g.finished:
            return
        withdrawn = [p for p in g.players if p.withdrawn]
        if not withdrawn:
            QMessageBox.information(self, "复赛", "当前没有退赛的选手。")
            return
        dlg = SelectPlayersDialog("复赛 — 勾选要恢复参赛的选手", withdrawn, self)
        if dlg.exec() != QDialog.Accepted:
            return
        nums = dlg.selected_numbers()
        if not nums:
            return
        if QMessageBox.question(self, "复赛确认",
                                "确定让选中的 %d 名选手复赛?\n当前轮将自动重排。" % len(nums),
                                QMessageBox.Yes | QMessageBox.No, QMessageBox.No) != QMessageBox.Yes:
            return
        g.reenter(nums)
        self._after_roster_change("已复赛 %d 人,当前轮已重排" % len(nums))

    def _after_roster_change(self, msg):
        """退赛/复赛后:重排了当前轮 → 重建该轮选项卡、切回对阵页。"""
        g = self.group
        self.rebuild_round_tab(g.current_round)
        self.tabs.setCurrentIndex(g.current_round - 1)
        self.stack.setCurrentWidget(self.tabs)
        self.refresh_status(); self.persist()
        self.statusBar().showMessage(msg, 3000)

    def after_roster_edit(self):
        """花名册改动后:存盘 + 刷新状态栏人数。"""
        self.persist()
        self.refresh_status()

    def on_roster_add(self):
        self.group.add_player("新选手")
        self.roster_table.reload()
        self.after_roster_edit()

    def on_roster_del(self):
        nums = self.roster_table.selected_numbers()
        for n in nums:
            self.group.remove_player(n)
        self.roster_table.reload()
        self.after_roster_edit()

    def on_roster_import_txt(self):
        fn, _ = QFileDialog.getOpenFileName(self, "导入名单(每行一个姓名)", "",
                                            "文本文件 (*.txt);;所有文件 (*)")
        if not fn:
            return
        n = 0
        with open(fn, encoding="utf-8") as f:
            for line in f:
                name = line.strip().split()[0] if line.strip() else ""
                if name:
                    self.group.add_player(name); n += 1
        self.roster_table.reload()
        self.after_roster_edit()
        self.statusBar().showMessage("已从 txt 导入 %d 人" % n, 4000)

    def on_roster_import_excel(self):
        fn, _ = QFileDialog.getOpenFileName(self, "选择 Excel 名单", "",
                                            "Excel 文件 (*.xlsx);;所有文件 (*)")
        if not fn:
            return
        try:
            dlg = ExcelImportDialog(fn, self)
        except Exception as e:
            QMessageBox.critical(self, "打开 Excel 失败", str(e))
            return
        if dlg.exec() != QDialog.Accepted:
            return
        players = dlg.players()
        if not players:
            QMessageBox.warning(self, "无数据", "没有读到有效的姓名行。")
            return
        for d in players:
            name = d.pop("name")
            self.group.add_player(name, **d)
        self.roster_table.reload()
        self.after_roster_edit()
        self.statusBar().showMessage("已从 Excel 导入 %d 人" % len(players), 4000)

    def on_draw_lots(self):
        if len(self.group.players) < 2:
            QMessageBox.warning(self, "人数不足", "至少需要 2 名选手才能开赛。")
            return
        if QMessageBox.question(self, "抽签开赛",
                                "将对 %d 名选手抽签并编排首轮。开赛后名单锁定(只能退赛/复赛)。\n确定开赛吗?"
                                % len(self.group.players),
                                QMessageBox.Yes | QMessageBox.No, QMessageBox.No) != QMessageBox.Yes:
            return
        self.group.start(same_team_avoid=self.cb_same_team.isChecked(),
                         lower_number_red=self.cb_lower_red.isChecked())
        self.set_active(self.cur_event, self.current_path, self.gi)
        self.statusBar().showMessage("已开赛,编排首轮 %d 台" % len(self.group.boards), 4000)

    def refresh_tree(self):
        """左栏只显示当前赛事(及其组别);切换不同赛事统一走工具栏「切换赛事」下拉。"""
        evs = tournament.list_events_with_groups()
        cur = [t for t in evs if t[1] == self.current_path]
        self.left.set_tree(cur, self.current_path, self.gi)

    def _populate_switch_menu(self):
        """工具栏「切换赛事」下拉:列出全部赛事。"""
        self.switch_menu.clear()
        for name, path in tournament.list_events():
            act = self.switch_menu.addAction(name, lambda checked=False, p=path: self.open_target(p, 0))
            act.setCheckable(True)
            act.setChecked(path == self.current_path)

    def open_target(self, path, gi):
        gi = 0 if gi is None else gi
        if path == self.current_path and gi == self.gi:
            return                                           # 已是当前组,不重复切换
        if path == self.current_path:
            self.set_active(self.cur_event, path, gi)        # 同赛事:仅切组
        else:
            self.set_active(tournament.load_event(path), path, gi)

    def persist(self):
        if self.cur_event and self.current_path:
            tournament.save_event(self.cur_event, self.current_path)

    # ── 选项卡 ────────────────────────────────────────────────────
    def build_all_tabs(self):
        g = self.group
        self.tabs.clear()
        total = max(g.total_rounds, g.current_round, 1)
        for r in range(1, total + 1):
            if r <= g.current_round:
                locked = (r != g.current_round) or g.finished   # 结算后末轮也锁定
                view = make_round_view(g.rounds[r - 1], self.after_score, locked)
            else:
                view = make_round_view([])
            self.tabs.addTab(view, round_label(r))
        self.tabs.setCurrentIndex(max(g.current_round - 1, 0))

    def rebuild_round_tab(self, r):
        g = self.group
        idx = r - 1
        if r <= g.current_round:
            view = make_round_view(g.rounds[r - 1], self.after_score, r != g.current_round)
        else:
            view = make_round_view([])
        self.tabs.removeTab(idx)
        self.tabs.insertTab(idx, view, round_label(r))

    # ── 状态栏 ────────────────────────────────────────────────────
    def refresh_status(self):
        g = self.group
        self.left.update_reports(g.started, g.finished)   # 报表按钮按状态启停
        if not g.started:                        # 报名阶段
            self.cur_event_lbl.setText("赛事:%s  ›  %s   |   报名阶段:%d 人"
                                       % (self.cur_event.name, g.name, len(g.players)))
            self.done_lbl.setText("未开赛")
            self.next_act.setEnabled(False)
            self.undo_act.setEnabled(False)
            self.hint_lbl.setText("")
            return
        self.cur_event_lbl.setText("赛事:%s  ›  %s   |   %s / 共 %d 轮"
                               % (self.cur_event.name, g.name, round_label(g.current_round), g.total_rounds))
        if g.finished:                            # 已结算
            self.done_lbl.setText("比赛已结束")
            self.next_act.setText("比赛已结束"); self.next_act.setEnabled(False)
            self.undo_act.setText("撤销结算 ⟲"); self.undo_act.setEnabled(True)
            self.hint_lbl.setText("")
            return
        self.done_lbl.setText("本轮完成度 %d / %d" % (g.done_count, len(g.boards)))
        # 末轮 → 计算最终成绩;否则 → 编排下一轮(登满才可用)
        self.next_act.setText("编排下一轮 ▶" if g.has_next_round else "计算最终成绩 ⚑")
        self.next_act.setEnabled(g.all_scored)
        # 首轮 → 重新抽签;第2轮起 → 撤销本轮(都可逆)
        self.undo_act.setText("◀ 撤销本轮" if g.current_round >= 2 else "重新抽签 ⟳")
        self.undo_act.setEnabled(g.current_round >= 1)
        unscored = [b.tai for b in g.boards if not b.is_bye and b.result == 0]
        if 0 < len(unscored) < 5:
            self.hint_lbl.setText("未登分:" + " ".join("第%d台" % n for n in unscored))
        else:
            self.hint_lbl.setText("")

    def after_score(self):
        self.refresh_status()
        self.persist()

    # ── 轮次推进 / 撤销 ───────────────────────────────────────────
    def on_next_round(self):
        g = self.group
        if not g.all_scored:
            return
        if not g.has_next_round:                  # 末轮:计算最终成绩、结束比赛
            return self.on_finish()
        prev = g.current_round
        g.record_and_advance()
        self.rebuild_round_tab(prev)
        self.rebuild_round_tab(g.current_round)
        self.tabs.setCurrentIndex(g.current_round - 1)
        self.refresh_status(); self.persist()
        self.statusBar().showMessage("已编排%s,共 %d 台" % (round_label(g.current_round), len(g.boards)), 3000)

    def on_finish(self):
        g = self.group
        if g.finished or not g.all_scored:
            return
        if QMessageBox.question(self, "计算最终成绩",
                                "将计入最后一轮成绩、计算最终名次并结束比赛。\n(可随时「撤销结算」)\n确定吗?",
                                QMessageBox.Yes | QMessageBox.No, QMessageBox.No) != QMessageBox.Yes:
            return
        g.finish()
        self.build_all_tabs()                     # 重建:末轮也锁定
        self.refresh_status(); self.persist()
        StandingsDialog(g, "%s · %s" % (self.cur_event.name, g.name),
                        self.cur_event.judge, self).exec()    # 弹出最终名次

    def on_undo(self):
        g = self.group
        if g.finished:                            # 已结算:撤销结算
            g.unfinish()
            self.build_all_tabs()
            self.refresh_status(); self.persist()
            self.statusBar().showMessage("已撤销结算,可继续修改", 3000)
            return
        if g.current_round == 1:                  # 首轮:重新抽签
            return self.on_redraw()
        if g.current_round < 1:
            return
        ans = QMessageBox.question(
            self, "撤销本轮",
            "将丢弃当前【%s】的全部登分,退回【%s】重新登分。\n确定撤销吗?"
            % (round_label(g.current_round), round_label(g.current_round - 1)),
            QMessageBox.Yes | QMessageBox.No, QMessageBox.No)
        if ans != QMessageBox.Yes:
            return
        discarded = g.current_round
        g.undo_last_round()
        self.rebuild_round_tab(discarded)
        self.rebuild_round_tab(g.current_round)
        self.tabs.setCurrentIndex(g.current_round - 1)
        self.refresh_status(); self.persist()
        self.statusBar().showMessage("已退回%s,可重新登分" % round_label(g.current_round), 3000)

    def on_redraw(self):
        g = self.group
        if g.current_round != 1:
            return
        if QMessageBox.question(self, "重新抽签",
                                "将丢弃当前首轮对阵(及已登分),重新洗签号并重排首轮。\n确定重新抽签吗?",
                                QMessageBox.Yes | QMessageBox.No, QMessageBox.No) != QMessageBox.Yes:
            return
        g.redraw_first_round()
        self.rebuild_round_tab(1)
        self.tabs.setCurrentIndex(0)
        self.refresh_status(); self.persist()
        self.statusBar().showMessage("已重新抽签,首轮 %d 台" % len(g.boards), 3000)

    # ── 赛事 / 组别管理 ───────────────────────────────────────────
    def on_open(self):
        target = self.left.selected_target()
        if not target:
            QMessageBox.information(self, "打开", "请先在左侧选择赛事或组别。")
            return
        path, gi = target
        self.open_target(path, gi)

    def on_new_event(self):
        dlg = NewEventDialog(self)
        if dlg.exec() != QDialog.Accepted:
            return
        name, judge, gname, rounds, names = dlg.data()
        if not name or not gname:
            QMessageBox.warning(self, "信息不全", "请填写赛事名称和组别名称(名单可在报名阶段再补录)。")
            return
        path = tournament.event_path(name)
        if os.path.exists(path):
            QMessageBox.warning(self, "赛事已存在", "已存在同名赛事,请改名。")
            return
        e = tournament.Event(name, judge=judge)
        e.add_group(tournament.Group.from_roster(gname, names, total_rounds=rounds))
        self.set_active(e, path, 0)
        self.statusBar().showMessage("已新建赛事:%s / %s(%d 人)" % (name, gname, len(names)), 4000)

    def on_new_group(self):
        if not self.cur_event:
            return
        dlg = NewGroupDialog(self)
        if dlg.exec() != QDialog.Accepted:
            return
        gname, rounds, names = dlg.data()
        if not gname:
            QMessageBox.warning(self, "信息不全", "请填写组别名称(名单可在报名阶段再补录)。")
            return
        if gname in self.cur_event.group_names():
            QMessageBox.warning(self, "组别已存在", "本赛事已有同名组别,请改名。")
            return
        self.cur_event.add_group(tournament.Group.from_roster(gname, names, total_rounds=rounds))
        self.set_active(self.cur_event, self.current_path, len(self.cur_event.groups) - 1)
        self.statusBar().showMessage("已新建组别:%s(%d 人)" % (gname, len(names)), 4000)

    def on_del_group(self):
        if not self.cur_event or len(self.cur_event.groups) < 2:
            QMessageBox.information(self, "删除组别", "每个赛事至少保留一个组别。\n如需删除整个赛事,请用「删除赛事」。")
            return
        gname = self.group.name
        if QMessageBox.question(self, "删除组别", "确定删除组别【%s】?此操作不可恢复。" % gname,
                                QMessageBox.Yes | QMessageBox.No, QMessageBox.No) != QMessageBox.Yes:
            return
        self.cur_event.remove_group(self.gi)
        self.set_active(self.cur_event, self.current_path, 0)
        self.statusBar().showMessage("已删除组别:%s" % gname, 3000)

    def on_edit_info(self):
        """编辑当前赛事信息:赛事名/裁判长 + 当前组别名/总轮数。"""
        if not self.cur_event:
            return
        dlg = EditInfoDialog(self.cur_event.name, self.cur_event.judge,
                             self.group.name, self.group.total_rounds, self)
        if dlg.exec() != QDialog.Accepted:
            return
        name, judge, gname, rounds = dlg.data()
        if not name or not gname:
            QMessageBox.warning(self, "信息不全", "赛事名称和组别名称不能为空。")
            return
        if gname in [g.name for i, g in enumerate(self.cur_event.groups) if i != self.gi]:
            QMessageBox.warning(self, "重名", "本赛事已有同名组别。")
            return
        new_path = self.current_path
        if name != self.cur_event.name:
            new_path = tournament.event_path(name)
            if os.path.exists(new_path):
                QMessageBox.warning(self, "重名", "已存在同名赛事,请改用其它名称。")
                return
        # 应用改动
        self.cur_event.judge = judge
        self.group.name = gname
        self.group.total_rounds = rounds
        if name != self.cur_event.name:
            self.cur_event.name = name
            tournament.save_event(self.cur_event, new_path)
            tournament.delete_event(self.current_path)
            self.current_path = new_path
        self.set_active(self.cur_event, self.current_path, self.gi)
        self.statusBar().showMessage("赛事信息已更新", 3000)

    def on_rename_event(self):
        old = self.cur_event.name
        new, ok = QInputDialog.getText(self, "重命名赛事", "新赛事名称:", text=old)
        new = new.strip()
        if not ok or not new or new == old:
            return
        new_path = tournament.event_path(new)
        if os.path.exists(new_path):
            QMessageBox.warning(self, "重名", "已存在同名赛事,请改用其它名称。")
            return
        self.cur_event.name = new
        tournament.save_event(self.cur_event, new_path)   # 存到新文件名
        tournament.delete_event(self.current_path)        # 删旧文件
        self.current_path = new_path
        self.set_active(self.cur_event, new_path, self.gi)
        self.statusBar().showMessage("赛事已改名:%s → %s" % (old, new), 4000)

    def on_rename_group(self):
        old = self.group.name
        new, ok = QInputDialog.getText(self, "重命名组别", "新组别名称:", text=old)
        new = new.strip()
        if not ok or not new or new == old:
            return
        if new in [g.name for i, g in enumerate(self.cur_event.groups) if i != self.gi]:
            QMessageBox.warning(self, "重名", "本赛事已有同名组别,请改用其它名称。")
            return
        self.group.name = new
        self.set_active(self.cur_event, self.current_path, self.gi)
        self.statusBar().showMessage("组别已改名:%s → %s" % (old, new), 4000)

    def on_del_event(self, path=None):
        if not path:                                # 不带参数 → 用左侧选中项(右键菜单)
            target = self.left.selected_target()
            if not target:
                QMessageBox.information(self, "删除赛事", "请先在左侧选择赛事。")
                return
            path = target[0]
        try:
            name = tournament.load_event(path).name
        except Exception:
            name = path
        if QMessageBox.question(self, "删除赛事", "确定删除整个赛事【%s】(含全部组别)?不可恢复。" % name,
                                QMessageBox.Yes | QMessageBox.No, QMessageBox.No) != QMessageBox.Yes:
            return
        was_current = (path == self.current_path)
        tournament.delete_event(path)
        events = tournament.list_events()
        if was_current:
            if events:
                self.set_active(tournament.load_event(events[0][1]), events[0][1], 0)
            else:
                self.load_startup()
        else:
            self.refresh_tree()

    # ── 工具栏 / 报表 ─────────────────────────────────────────────
    def on_tool(self, text):
        if text == "随机结果(测试)":
            self.on_random_results()
        elif text == "花名册":
            self.on_roster()
        elif text == "生成秩序册":
            self.export_orderbook()
        else:
            self.statusBar().showMessage("工具栏:%s(功能待接入)" % text, 2000)

    def on_random_results(self):
        """开发测试用:给当前轮每台随机判一个结果,方便快速灌数据。"""
        import random
        g = self.group
        if not g.started or g.finished:
            return
        for b in g.boards:
            if not b.is_bye:
                b.result = random.choice([1, 2, 3])
        self.rebuild_round_tab(g.current_round)   # 重建当前轮,卡片显示结果
        self.tabs.setCurrentIndex(g.current_round - 1)
        self.refresh_status(); self.persist()
        self.statusBar().showMessage("已随机生成本轮结果(仅测试用)", 2000)

    def on_report(self, name):
        if name == "本轮对阵表":
            self.print_pairing()
        elif name == "总成绩表":
            if self.group.finished:
                StandingsDialog(self.group, "%s · %s" % (self.cur_event.name, self.group.name),
                                self.cur_event.judge, self).exec()
        elif name == "选手花名册":
            self.export_roster()
        elif name == "台号卡":
            self.export_tablecards()
        elif name == "荣誉证书":
            self.export_certificates()
        elif name == "记分卡":
            self.export_scorecards()
        else:
            self.statusBar().showMessage("报表:%s(待接入)" % name, 2000)

    @staticmethod
    def _orderbook_bg():
        base = os.path.dirname(os.path.abspath(__file__))
        for fn in ("秩序册封面.png", "秩序册封面.jpg", "秩序册封面.jpeg"):
            p = os.path.join(base, fn)
            if os.path.exists(p):
                return p
        return None

    def export_orderbook(self):
        """秩序册:我们出封面+封底,中间夹用户自撰的内容 PDF(可选)。"""
        import datetime
        ev = self.cur_event
        content, _ = QFileDialog.getOpenFileName(
            self, "选择秩序册内容 PDF(主办方自撰;取消则只出封面封底)", "",
            "PDF 文件 (*.pdf);;所有文件 (*)")
        out = os.path.join(tournament.events_dir(),
                           "%s_竞赛秩序册.pdf" % tournament.safe_filename(ev.name))
        try:
            import print_pdf
            print_pdf.generate_orderbook_pdf(
                ev.name, ev.judge, datetime.date.today().strftime("%Y年%m月%d日"),
                out, content_pdf=(content or None), cover_bg=self._orderbook_bg())
        except Exception as e:
            QMessageBox.critical(self, "生成 PDF 失败", str(e)); return
        open_pdf_or_inform(self, out)
        msg = "已生成秩序册(封面 + 你的内容 + 封底)" if content else "已生成秩序册封面+封底(未选内容)"
        self.statusBar().showMessage(msg, 4000)

    def export_scorecards(self):
        g = self.group
        if not g.started or not g.boards:
            QMessageBox.information(self, "记分卡", "请先开赛、编排出对阵后再生成记分卡。")
            return
        cards = []
        for b in g.boards:
            red = "%s(%d)" % (b.red.name, b.red.number)
            black = "轮空" if b.is_bye else "%s(%d)" % (b.black.name, b.black.number)
            cards.append({"tai": b.tai, "red": red, "black": black})
        title = "%s · %s · %s" % (self.cur_event.name, g.name, round_label(g.current_round))
        out = os.path.join(tournament.events_dir(), "%s_%s_%s记分卡.pdf"
                           % (tournament.safe_filename(self.cur_event.name),
                              tournament.safe_filename(g.name), round_label(g.current_round)))
        try:
            import print_pdf
            print_pdf.generate_scorecards_pdf(cards, title, out)
        except Exception as e:
            QMessageBox.critical(self, "生成 PDF 失败", str(e)); return
        open_pdf_or_inform(self, out)
        self.statusBar().showMessage("已生成记分卡 %d 张" % len(cards), 4000)

    @staticmethod
    def _certificate_bg():
        base = os.path.dirname(os.path.abspath(__file__))
        for fn in ("荣誉证书底图.png", "荣誉证书底图.jpg", "荣誉证书底图.jpeg"):
            p = os.path.join(base, fn)
            if os.path.exists(p):
                return p
        return None

    def export_certificates(self):
        import datetime
        g = self.group
        if not g.finished:
            QMessageBox.information(self, "荣誉证书", "请先「计算最终成绩」结束比赛后再生成证书。")
            return
        n, ok = QInputDialog.getInt(self, "荣誉证书", "为前几名生成证书?", 3, 1, len(g.players))
        if not ok:
            return
        winners = [(rank, p.name) for rank, p, note, hl in g.standings()[:n]]
        bg = self._certificate_bg()
        out = os.path.join(tournament.events_dir(), "%s_%s_荣誉证书.pdf"
                           % (tournament.safe_filename(self.cur_event.name), tournament.safe_filename(g.name)))
        try:
            import print_pdf
            print_pdf.generate_certificates_pdf(
                winners, self.cur_event.name, g.name, out, bg_image=bg,
                judge=self.cur_event.judge, date_str=datetime.date.today().strftime("%Y年%m月%d日"))
        except Exception as e:
            QMessageBox.critical(self, "生成 PDF 失败", str(e)); return
        open_pdf_or_inform(self, out)
        self.statusBar().showMessage("已生成前 %d 名荣誉证书%s" % (n, "" if bg else "(未找到底图,白底)"), 4000)

    @staticmethod
    def _tablecard_bg():
        """台号卡底图:程序目录下「台号卡底图.png/.jpg」存在则用,否则无底图。"""
        base = os.path.dirname(os.path.abspath(__file__))
        for fn in ("台号卡底图.png", "台号卡底图.jpg", "台号卡底图.jpeg"):
            p = os.path.join(base, fn)
            if os.path.exists(p):
                return p
        return None

    def export_tablecards(self):
        g = self.group
        if not g.started or not g.boards:
            QMessageBox.information(self, "台号卡", "请先开赛、编排出对阵后再生成台号卡。")
            return
        cards = []
        for b in g.boards:
            red = "%s(%d)" % (b.red.name, b.red.number)
            black = "轮空" if b.is_bye else "%s(%d)" % (b.black.name, b.black.number)
            cards.append({"tai": b.tai, "red": red, "black": black})
        bg = self._tablecard_bg()
        title = "%s · %s · %s" % (self.cur_event.name, g.name, round_label(g.current_round))
        out = os.path.join(tournament.events_dir(), "%s_%s_%s台号卡.pdf"
                           % (tournament.safe_filename(self.cur_event.name),
                              tournament.safe_filename(g.name), round_label(g.current_round)))
        try:
            import print_pdf
            print_pdf.generate_tablecards_pdf(cards, title, out, bg_image=bg)
        except Exception as e:
            QMessageBox.critical(self, "生成 PDF 失败", str(e)); return
        open_pdf_or_inform(self, out)
        if bg:
            self.statusBar().showMessage("已生成台号卡(已套用底图)", 4000)
        else:
            self.statusBar().showMessage("已生成台号卡(无底图;把图片命名为「台号卡底图.png」放到程序目录即可套用)", 6000)

    def export_roster(self):
        g = self.group
        if not g.players:
            QMessageBox.information(self, "选手花名册", "当前没有选手。")
            return
        rows = []
        for p in sorted(g.players, key=lambda x: x.number):
            name = p.name + ("(退赛)" if p.withdrawn else "")
            rows.append([str(p.number), name, p.gender, p.team, p.idcard, p.phone, str(p.rating)])
        title = "%s · %s" % (self.cur_event.name, g.name)
        out = os.path.join(tournament.events_dir(), "%s_%s_选手花名册.pdf"
                           % (tournament.safe_filename(self.cur_event.name), tournament.safe_filename(g.name)))
        try:
            import print_pdf
            print_pdf.generate_roster_pdf(rows, title, out, judge=self.cur_event.judge)
        except Exception as e:
            QMessageBox.critical(self, "生成 PDF 失败", str(e)); return
        open_pdf_or_inform(self, out)
        self.statusBar().showMessage("已生成:%s" % out, 4000)

    def print_pairing(self):
        try:
            import print_pdf
            g = self.group
            title = "%s · %s" % (self.cur_event.name, g.name)
            out = os.path.join(tournament.events_dir(), "%s_%s_%s对阵表.pdf"
                               % (tournament.safe_filename(self.cur_event.name),
                                  tournament.safe_filename(g.name), round_label(g.current_round)))
            print_pdf.generate_pairing_pdf(g.pdf_rows(), title, g.current_round, out, judge=self.cur_event.judge)
        except Exception as e:
            QMessageBox.critical(self, "生成 PDF 失败", str(e)); return
        open_pdf_or_inform(self, out)
        self.statusBar().showMessage("已生成:%s" % out, 4000)


def main():
    app = QApplication(sys.argv)
    app.setStyleSheet(STYLE)
    win = MainWindow()
    win.show()
    sys.exit(app.exec())


if __name__ == "__main__":
    main()
